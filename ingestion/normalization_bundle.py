"""Validated Excel bootstrap for a portable Transportstyrelsen normalization snapshot."""

from __future__ import annotations

import json
from dataclasses import dataclass, replace
from datetime import UTC, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from psycopg import Connection
from psycopg.types.json import Jsonb

from ingestion.job_bookkeeping_migrations import run_job_bookkeeping_migrations
from ingestion.normalization_migrations import (
    NORMALIZATION_RESULTS_TABLE,
    TRANSLATION_RULE_VERSIONS_TABLE,
    run_normalization_migrations,
)
from ingestion.normalization_rules import (
    MAPPING_VERSION,
    PIPELINE_VERSION,
    ManufacturerEntityRules,
    manufacturer_entity_catalog,
    normalize_manufacturer_entity,
)
from ingestion.normalization_service import normalize_batch
from ingestion.review_queue_migrations import run_review_queue_migrations
from ingestion.staging_migrations import run_staging_migrations
from ingestion.translation_dictionaries import (
    TranslationRuleSet,
    load_translation_rule_set,
)

SOURCE_TABLE = "staging.transportstyrelsen_raw"
REQUIRED_SHEETS = frozenset(
    {
        "TS Raw Records",
        "Normalized Results",
        "Translation Rules",
        "Base Manufacturer Catalog",
        "Effective Mfr Entities",
        "Manufacturer Overrides",
        "Policy Overrides",
        "Rule Version",
    }
)


class NormalizationBundleError(ValueError):
    """Raised when a workbook or target database cannot reproduce its snapshot."""


@dataclass(frozen=True)
class BundleRawRecord:
    staging_id: int
    source_batch_id: str
    ingested_at: datetime
    raw_record: dict[str, Any]


@dataclass(frozen=True)
class BundleExpectedResult:
    source_record_id: int
    source_batch_id: str
    mapping_version: str
    rule_version: str
    pipeline_version: str
    status: str
    normalized_payload: dict[str, Any]
    applied_rule_ids: list[str]
    review_reasons: list[str]
    confidence: float


@dataclass(frozen=True)
class BundleRuleVersion:
    version: str
    base_rule_version: str
    activation_note: str
    activated_at: datetime
    overrides: dict[str, dict[str, Any]]


@dataclass(frozen=True)
class NormalizationBundle:
    source_batch_id: str
    raw_records: tuple[BundleRawRecord, ...]
    expected_results: tuple[BundleExpectedResult, ...]
    rule_version: BundleRuleVersion
    translation_rule_count: int
    base_manufacturer_count: int
    effective_manufacturer_count: int
    manufacturer_override_count: int
    policy_override_count: int


@dataclass(frozen=True)
class BundleImportSummary:
    source_batch_id: str
    rule_version: str
    raw_records: int
    normalized_results: int
    resolved: int
    provisional: int
    review_required: int
    failed: int
    verified: bool


def _sheet_rows(workbook: Any, sheet_name: str) -> tuple[dict[str, Any], ...]:
    sheet = workbook[sheet_name]
    headers = tuple(
        str(cell.value).strip() if cell.value is not None else "" for cell in sheet[4]
    )
    if not headers or any(not header for header in headers):
        raise NormalizationBundleError(f"{sheet_name} has an invalid row-4 header")
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=5, values_only=True):
        if all(value in (None, "") for value in values):
            continue
        rows.append(dict(zip(headers, values, strict=False)))
    return tuple(rows)


def _required(row: dict[str, Any], column: str, *, sheet: str) -> Any:
    value = row.get(column)
    if value in (None, ""):
        raise NormalizationBundleError(f"{sheet}.{column} must not be empty")
    return value


def _json_value(
    row: dict[str, Any],
    column: str,
    *,
    sheet: str,
    expected_type: type[Any],
) -> Any:
    value = _required(row, column, sheet=sheet)
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError as error:
        raise NormalizationBundleError(f"{sheet}.{column} contains invalid JSON") from error
    if not isinstance(parsed, expected_type):
        raise NormalizationBundleError(
            f"{sheet}.{column} must contain a JSON {expected_type.__name__}"
        )
    return parsed


def _datetime_value(value: Any, *, field: str) -> datetime:
    if isinstance(value, datetime):
        return value if value.tzinfo is not None else value.replace(tzinfo=UTC)
    if isinstance(value, (int, float)):
        parsed = from_excel(value)
        if not isinstance(parsed, datetime):
            raise NormalizationBundleError(f"{field} must contain a timestamp")
        return parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=UTC)
    try:
        return datetime.fromisoformat(str(value))
    except ValueError as error:
        raise NormalizationBundleError(f"{field} must contain an ISO timestamp") from error


def _definition_map(
    rows: tuple[dict[str, Any], ...],
    *,
    sheet: str,
    id_column: str,
    json_column: str,
) -> dict[str, dict[str, Any]]:
    definitions: dict[str, dict[str, Any]] = {}
    for row in rows:
        identifier = str(_required(row, id_column, sheet=sheet))
        if identifier in definitions:
            raise NormalizationBundleError(f"{sheet} contains duplicate ID {identifier}")
        definitions[identifier] = _json_value(
            row,
            json_column,
            sheet=sheet,
            expected_type=dict,
        )
    return definitions


def load_normalization_bundle(path: str | Path) -> NormalizationBundle:
    """Read and validate the complete workbook contract without touching a database."""

    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise NormalizationBundleError("normalization bundle file does not exist")
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as error:
        raise NormalizationBundleError("normalization bundle is not a readable XLSX file") from error
    try:
        missing_sheets = REQUIRED_SHEETS.difference(workbook.sheetnames)
        if missing_sheets:
            missing = ", ".join(sorted(missing_sheets))
            raise NormalizationBundleError(f"normalization bundle is missing sheets: {missing}")

        raw_rows = _sheet_rows(workbook, "TS Raw Records")
        result_rows = _sheet_rows(workbook, "Normalized Results")
        version_rows = _sheet_rows(workbook, "Rule Version")
        if not raw_rows or not result_rows:
            raise NormalizationBundleError("normalization bundle must contain source and result rows")
        if len(version_rows) != 1:
            raise NormalizationBundleError("Rule Version must contain exactly one row")

        raw_records = tuple(
            BundleRawRecord(
                staging_id=int(_required(row, "staging_id", sheet="TS Raw Records")),
                source_batch_id=str(
                    _required(row, "source_batch_id", sheet="TS Raw Records")
                ),
                ingested_at=_datetime_value(
                    _json_value(
                        row,
                        "ingested_at_json",
                        sheet="TS Raw Records",
                        expected_type=str,
                    ),
                    field="TS Raw Records.ingested_at_json",
                ),
                raw_record=_json_value(
                    row,
                    "raw_record_json",
                    sheet="TS Raw Records",
                    expected_type=dict,
                ),
            )
            for row in raw_rows
        )
        expected_results = tuple(
            BundleExpectedResult(
                source_record_id=int(
                    _required(row, "source_record_id", sheet="Normalized Results")
                ),
                source_batch_id=str(
                    _required(row, "source_batch_id", sheet="Normalized Results")
                ),
                mapping_version=str(
                    _required(row, "mapping_version", sheet="Normalized Results")
                ),
                rule_version=str(
                    _required(row, "rule_version", sheet="Normalized Results")
                ),
                pipeline_version=str(
                    _required(row, "pipeline_version", sheet="Normalized Results")
                ),
                status=str(_required(row, "status", sheet="Normalized Results")),
                normalized_payload=_json_value(
                    row,
                    "normalized_payload_json",
                    sheet="Normalized Results",
                    expected_type=dict,
                ),
                applied_rule_ids=_json_value(
                    row,
                    "applied_rule_ids",
                    sheet="Normalized Results",
                    expected_type=list,
                ),
                review_reasons=_json_value(
                    row,
                    "review_reasons",
                    sheet="Normalized Results",
                    expected_type=list,
                ),
                confidence=float(
                    _required(row, "confidence", sheet="Normalized Results")
                ),
            )
            for row in result_rows
        )

        version_row = version_rows[0]
        rule_version = BundleRuleVersion(
            version=str(_required(version_row, "version", sheet="Rule Version")),
            base_rule_version=str(
                _required(version_row, "base_rule_version", sheet="Rule Version")
            ),
            activation_note=str(
                _required(version_row, "activation_note", sheet="Rule Version")
            ),
            activated_at=_datetime_value(
                _json_value(
                    version_row,
                    "activated_at_json",
                    sheet="Rule Version",
                    expected_type=str,
                ),
                field="Rule Version.activated_at_json",
            ),
            overrides=_json_value(
                version_row,
                "overrides_json",
                sheet="Rule Version",
                expected_type=dict,
            ),
        )
        try:
            load_translation_rule_set(rule_version.base_rule_version)
        except LookupError as error:
            raise NormalizationBundleError(
                "bundle base rule version is unavailable in this application"
            ) from error

        batch_ids = {record.source_batch_id for record in raw_records}
        if len(batch_ids) != 1:
            raise NormalizationBundleError("TS Raw Records must contain one source batch")
        source_batch_id = next(iter(batch_ids))
        raw_ids = {record.staging_id for record in raw_records}
        result_ids = {result.source_record_id for result in expected_results}
        if len(raw_ids) != len(raw_records) or len(result_ids) != len(expected_results):
            raise NormalizationBundleError("source and result record IDs must be unique")
        if raw_ids != result_ids:
            raise NormalizationBundleError("source and expected result IDs do not match")
        if any(result.source_batch_id != source_batch_id for result in expected_results):
            raise NormalizationBundleError("expected results reference a different source batch")
        if any(result.mapping_version != MAPPING_VERSION for result in expected_results):
            raise NormalizationBundleError("expected results use an incompatible mapping version")
        if any(result.pipeline_version != PIPELINE_VERSION for result in expected_results):
            raise NormalizationBundleError("expected results use an incompatible pipeline version")
        if any(result.rule_version != rule_version.version for result in expected_results):
            raise NormalizationBundleError("expected results use a different rule version")

        manufacturer_rows = _sheet_rows(workbook, "Manufacturer Overrides")
        policy_rows = _sheet_rows(workbook, "Policy Overrides")
        manufacturer_definitions = _definition_map(
            manufacturer_rows,
            sheet="Manufacturer Overrides",
            id_column="rule_id",
            json_column="definition_json",
        )
        policy_definitions = _definition_map(
            policy_rows,
            sheet="Policy Overrides",
            id_column="rule_id",
            json_column="definition_json",
        )
        expected_manufacturers = {
            key: value
            for key, value in rule_version.overrides.items()
            if value.get("kind") == "manufacturer_entity"
        }
        expected_policies = {
            key: value
            for key, value in rule_version.overrides.items()
            if value.get("kind") != "manufacturer_entity"
        }
        if manufacturer_definitions != expected_manufacturers:
            raise NormalizationBundleError(
                "Manufacturer Overrides does not match Rule Version.overrides_json"
            )
        if policy_definitions != expected_policies:
            raise NormalizationBundleError(
                "Policy Overrides does not match Rule Version.overrides_json"
            )

        translation_rows = _sheet_rows(workbook, "Translation Rules")
        translation_definitions = _definition_map(
            translation_rows,
            sheet="Translation Rules",
            id_column="rule_id",
            json_column="effective_rule_json",
        )
        base_rule_set = load_translation_rule_set(rule_version.base_rule_version)
        if set(translation_definitions) != set(base_rule_set.by_id):
            raise NormalizationBundleError(
                "Translation Rules does not match the application base catalog"
            )
        for rule in base_rule_set.rules:
            definition = translation_definitions[rule.rule_id]
            override = rule_version.overrides.get(rule.rule_id, {})
            actual = (
                definition.get("area"),
                tuple(definition.get("source_fields", [])),
                tuple(definition.get("source_terms", [])),
                definition.get("canonical_field"),
                definition.get(
                    "effective_canonical_value",
                    definition.get("canonical_value"),
                ),
                definition.get("effective_decision", definition.get("decision")),
                definition.get(
                    "effective_display_value",
                    definition.get("display_value"),
                ),
                tuple(definition.get("vehicle_scopes", [])),
                tuple(definition.get("manufacturers", [])),
            )
            expected = (
                rule.area,
                rule.source_fields,
                rule.source_terms,
                rule.canonical_field,
                override.get("canonical_value", rule.canonical_value),
                override.get("decision", rule.decision),
                override.get("display_value", rule.display_value),
                rule.vehicle_scopes,
                rule.manufacturers,
            )
            if actual != expected:
                raise NormalizationBundleError(
                    f"Translation Rules definition {rule.rule_id} is incompatible"
                )

        base_manufacturer_rows = _sheet_rows(workbook, "Base Manufacturer Catalog")
        base_definitions = _definition_map(
            base_manufacturer_rows,
            sheet="Base Manufacturer Catalog",
            id_column="entity_id",
            json_column="base_definition_json",
        )
        expected_base_definitions: dict[str, dict[str, Any]] = {}
        for catalog_definition in manufacturer_entity_catalog():
            source_field = str(catalog_definition["source_field"])
            source_term = str(catalog_definition["source_term"])
            digest = sha256(f"{source_field}:{source_term}".encode()).hexdigest()
            expected_base_definitions[f"MFE-{digest[:14].upper()}"] = dict(
                catalog_definition
            )
        if base_definitions != expected_base_definitions:
            raise NormalizationBundleError(
                "Base Manufacturer Catalog does not match the application catalog"
            )
        effective_manufacturer_rows = _sheet_rows(workbook, "Effective Mfr Entities")
        effective_ids = {
            str(_required(row, "entity_id", sheet="Effective Mfr Entities"))
            for row in effective_manufacturer_rows
        }
        if len(effective_ids) != len(effective_manufacturer_rows):
            raise NormalizationBundleError("Effective Mfr Entities contains duplicate IDs")

        return NormalizationBundle(
            source_batch_id=source_batch_id,
            raw_records=raw_records,
            expected_results=expected_results,
            rule_version=rule_version,
            translation_rule_count=len(translation_rows),
            base_manufacturer_count=len(base_manufacturer_rows),
            effective_manufacturer_count=len(effective_manufacturer_rows),
            manufacturer_override_count=len(manufacturer_rows),
            policy_override_count=len(policy_rows),
        )
    finally:
        workbook.close()


def _effective_rule_set(bundle: NormalizationBundle) -> TranslationRuleSet:
    base = load_translation_rule_set(bundle.rule_version.base_rule_version)
    rules = []
    for rule in base.rules:
        override = bundle.rule_version.overrides.get(rule.rule_id)
        if override is None:
            rules.append(rule)
            continue
        if override.get("kind") is not None:
            raise NormalizationBundleError(
                f"translation override {rule.rule_id} has an incompatible kind"
            )
        rules.append(
            replace(
                rule,
                canonical_value=override.get("canonical_value"),
                decision=override["decision"],
                display_value=override.get("display_value"),
            )
        )
    return TranslationRuleSet(version=bundle.rule_version.version, rules=tuple(rules))


def _manufacturer_rules(bundle: NormalizationBundle) -> ManufacturerEntityRules:
    rules: dict[str, dict[str, Any]] = {}
    for entity_id, override in bundle.rule_version.overrides.items():
        if override.get("kind") in {
            "manufacturer_match_policy",
            "special_vehicle_policy",
        }:
            rules[f"policy:{entity_id}"] = dict(override)
            continue
        if override.get("kind") != "manufacturer_entity":
            continue
        source_term = normalize_manufacturer_entity(override.get("source_term"))
        source_field = override.get("source_field")
        if source_term is None or not isinstance(source_field, str):
            raise NormalizationBundleError(f"manufacturer override {entity_id} is invalid")
        rules[f"{source_field}:{source_term}"] = {
            "entity_id": entity_id,
            "kind": "manufacturer_entity",
            "source_field": source_field,
            "source_term": source_term,
            "canonical_name": override.get("canonical_name"),
            "entity_role": override.get("entity_role"),
            "base_behavior": override.get("base_behavior"),
            "match_type": override.get("match_type"),
            "reviewed_examples": list(override.get("reviewed_examples") or []),
            "aliases": list(override.get("aliases") or []),
            "marketed_brand_overrides": dict(
                override.get("marketed_brand_overrides") or {}
            ),
            "fallback_manufacturer": override.get("fallback_manufacturer"),
        }
    return rules


def _store_rule_version(connection: Connection, bundle: NormalizationBundle) -> None:
    version = bundle.rule_version
    with connection.cursor() as cursor:
        cursor.execute(
            f"INSERT INTO {TRANSLATION_RULE_VERSIONS_TABLE} "
            "(version, base_rule_version, overrides, activation_note, activated_at) "
            "VALUES (%s, %s, %s, %s, %s) ON CONFLICT (version) DO NOTHING",
            (
                version.version,
                version.base_rule_version,
                Jsonb(version.overrides),
                version.activation_note,
                version.activated_at,
            ),
        )
        cursor.execute(
            f"SELECT base_rule_version, overrides, activation_note, activated_at "
            f"FROM {TRANSLATION_RULE_VERSIONS_TABLE} WHERE version = %s",
            (version.version,),
        )
        row = cursor.fetchone()
    if row is None:
        raise NormalizationBundleError("rule version insert returned no row")
    actual = (str(row[0]), dict(row[1]), str(row[2]), row[3])
    expected = (
        version.base_rule_version,
        version.overrides,
        version.activation_note,
        version.activated_at,
    )
    if actual != expected:
        raise NormalizationBundleError("target database contains a different rule version")


def _store_raw_records(connection: Connection, bundle: NormalizationBundle) -> None:
    with connection.cursor() as cursor:
        for record in bundle.raw_records:
            cursor.execute(
                f"INSERT INTO {SOURCE_TABLE} (id, source_batch_id, ingested_at, raw_record) "
                "VALUES (%s, %s, %s, %s) ON CONFLICT (id) DO NOTHING",
                (
                    record.staging_id,
                    record.source_batch_id,
                    record.ingested_at,
                    Jsonb(record.raw_record),
                ),
            )
        cursor.execute(
            f"SELECT id, source_batch_id, ingested_at, raw_record FROM {SOURCE_TABLE} "
            "WHERE source_batch_id = %s ORDER BY id",
            (bundle.source_batch_id,),
        )
        rows = cursor.fetchall()
        cursor.execute(
            "SELECT setval(pg_get_serial_sequence(%s, 'id'), "
            f"greatest((SELECT coalesce(max(id), 1) FROM {SOURCE_TABLE}), 1), true)",
            (SOURCE_TABLE,),
        )
    actual = tuple(
        (int(row[0]), str(row[1]), row[2], dict(row[3])) for row in rows
    )
    expected = tuple(
        (
            record.staging_id,
            record.source_batch_id,
            record.ingested_at,
            record.raw_record,
        )
        for record in bundle.raw_records
    )
    if actual != expected:
        raise NormalizationBundleError(
            "target database source batch differs from the workbook"
        )


def _verify_results(connection: Connection, bundle: NormalizationBundle) -> None:
    with connection.cursor() as cursor:
        cursor.execute(
            f"SELECT source_record_id, source_batch_id, mapping_version, rule_version, "
            "pipeline_version, status, normalized_payload, applied_rule_ids, "
            f"review_reasons, confidence FROM {NORMALIZATION_RESULTS_TABLE} "
            "WHERE source_batch_id = %s ORDER BY source_record_id",
            (bundle.source_batch_id,),
        )
        rows = cursor.fetchall()
    actual = tuple(
        (
            int(row[0]),
            str(row[1]),
            str(row[2]),
            str(row[3]),
            str(row[4]),
            str(row[5]),
                dict(row[6]),
                list(row[7]),
                list(row[8]),
            float(row[9]),
        )
        for row in rows
    )
    expected = tuple(
        (
            result.source_record_id,
            result.source_batch_id,
            result.mapping_version,
            result.rule_version,
            result.pipeline_version,
            result.status,
            result.normalized_payload,
            result.applied_rule_ids,
            result.review_reasons,
            result.confidence,
        )
        for result in bundle.expected_results
    )
    if actual != expected:
        mismatch_ids = [
            result.source_record_id
            for index, result in enumerate(bundle.expected_results)
            if index >= len(actual) or actual[index] != expected[index]
        ][:10]
        mismatch_fields = []
        if actual and expected:
            field_names = ("source_record_id", "source_batch_id", "mapping_version", "rule_version", "pipeline_version", "status", "normalized_payload", "applied_rule_ids", "review_reasons", "confidence")
            mismatch_fields = [name for name, left, right in zip(field_names, actual[0], expected[0]) if left != right]
            if "normalized_payload" in mismatch_fields:
                actual_payload = actual[0][6]
                expected_payload = expected[0][6]
                payload_differences = [
                    key for key in sorted(set(actual_payload) | set(expected_payload))
                    if actual_payload.get(key) != expected_payload.get(key)
                ]
                mismatch_fields.append(f"payload_keys={payload_differences}")
        raise NormalizationBundleError(
            "normalized results differ from the workbook"
            + (f" for source IDs {mismatch_ids}" if mismatch_ids else "")
            + (f"; differing fields: {mismatch_fields}" if mismatch_fields else "")
        )


def import_normalization_bundle(
    connection: Connection,
    path: str | Path,
) -> BundleImportSummary:
    """Populate PostgreSQL, run normalization, and prove it matches the workbook."""

    bundle = load_normalization_bundle(path)
    run_staging_migrations(connection)
    run_review_queue_migrations(connection)
    run_job_bookkeeping_migrations(connection)
    run_normalization_migrations(connection)
    try:
        _store_rule_version(connection, bundle)
        _store_raw_records(connection, bundle)
        connection.commit()
    except Exception:
        connection.rollback()
        raise

    summary = normalize_batch(
        connection,
        batch_id=bundle.source_batch_id,
        rule_set=_effective_rule_set(bundle),
        manufacturer_entity_rules=_manufacturer_rules(bundle),
    )
    _verify_results(connection, bundle)
    return BundleImportSummary(
        source_batch_id=bundle.source_batch_id,
        rule_version=bundle.rule_version.version,
        raw_records=len(bundle.raw_records),
        normalized_results=summary.processed,
        resolved=summary.resolved,
        provisional=summary.provisional,
        review_required=summary.review_required,
        failed=summary.failed,
        verified=True,
    )
